#NOBINS.PY

'''
	This script reads the CEDNet data files and turns out an excel spreadsheet with all products
	that don't have a bin location in the system.  It also gives you a sheet with all
	of the products in order for you to help punch them in using ShelfAddresses.py

'''
from openpyxl import Workbook

from openpyxl.styles import Side, Border
import CEDNetUtils as CED

CELL_FORMULA = "=IF(AND(G{s}<>\"\", H{s}<>\"\"), UPPER(CONCATENATE(G{s}, \".\"" \
					",H{s}, IF(I{s}<>\"\", CONCATENATE(\".\", I{s}),\"\"))),\"\")"
def run(skip_prompt=False):

	#The only time it will be skipped is on setup.  
	if not skip_prompt:
		while True:
			print("This will overwrite the current stock status workbook\nDo you wish to continue? Y/n")
			choice = input("??")
			if choice.lower() == "y":
				break
			if choice.lower() == "n":
				return
			print("Invalid choice.")

			
	product_list = CED.get_products()

	
	#Create and size the workbook
	workbook = Workbook()
	
	worksheet = workbook.get_sheet_by_name("Sheet")
	worksheet.column_dimensions['A'].width = 8
	worksheet.column_dimensions['B'].width = 24
	worksheet.column_dimensions['C'].width = 40
	worksheet.column_dimensions['D'].width = 10
	worksheet.column_dimensions['E'].width = 10
	side = Side(border_style='thin', color="00000000")
	
	border = Border(left=side,right=side,bottom=side, top=side)


	header = ["MFR", "CAT #", "DESCRIPTION", "OH Qty.", "Current Bin", "New Bin"]
	
	#write the excel sheet with all products
	worksheet = CED.write_excel_sheet(header = header, list = product_list, sheet = worksheet, border = border)

	nobinsheet = workbook.create_sheet('nobins', 1)

	
	#Read in a file with Manufacturers and Cat#'s that shouldn't be considered for bin locations.
	#Or create it if there isn't one
	dontcount = []
	try:
		f = open("C:\\PaulScripts\\Shelving Addresses\\na_bins.txt", 'r')
		for line in f:
			dontcount.append(line.strip())
			
	except FileNotFoundError:
		f=open("C:\\PaulScripts\\Shelving Addresses\\na_bins.txt", 'w')
		print("There is no file with things to exclude. \nI just made one at C:\\PaulScripts\\Shelving Addresses\\na_bins.txt")
		f.close
	
	if dontcount == []:
		print("You can add manufacturers and product numbers that shouldn't be assigned bin location\nto the file at C:\\PaulScripts\\Shelving Addresses\\na_bins.txt")
	else:	
		print("\n\n\n\n I will not include these\n__________________________")
		for each in dontcount:
			print(each)
		
		
	nobins = []
	for j, each in enumerate(product_list):

		if each[0].strip() in dontcount or each[1].strip() in dontcount:
		
			worksheet.cell(row=j+2, column=5).value = "n/a"
			continue

		if each[-1] == '' or each[-1] is None:

			nobins.append(each)


	header = ["MFR", "CAT #", "DESCRIPTION", "OH Qty.", "New Bin"]

	nobinsheet=CED.write_excel_sheet(header = header, list = nobins, sheet=nobinsheet, border = border)
	nobinsheet.column_dimensions['A'].width = 8
	nobinsheet.column_dimensions['B'].width = 24
	nobinsheet.column_dimensions['C'].width = 40
	nobinsheet.column_dimensions['D'].width = 10
	nobinsheet.column_dimensions['E'].width = 10
	
	
	while True:
		try:
			workbook.save("C:\\PaulScripts\\This Week's Stock Status.xlsx")
			break
		except:
			input("Please close the stock status excel document.  Press enter to continue")


	print("\n\nSpreadsheet saved to C:\\PaulScripts\\This Week's Stock Status.xlsx")
	
	print("\nThere are " + str(len(nobins)) + " items without bin locations.")

if __name__ == "__main__":
	run(skip_prompt=True)
else:
	print("imported nobins.py")
