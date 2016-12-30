#Solar Reporting


'''

     Reporting on solar sales is an arduous and boring process. 
     It needs to be run every week in various capacities.

    This script automates that process.

    There are three tiers to solar reporting: Weekly, monthly, and quarterly.
    They're all very similar, so a common process is not difficult to write.

    The process goes as follows:
        0.  On initialization, several global lists are populated with products that we need to report on.
            Also, instead of having tons of list accesses with cryptic numbers, I assigned a few common
            values to variables to make things easier to read.

            This script also requires a few files to be output from SPEAKS.  Just the sales for the
                 time period you'd like to report on, with customers sorted by job account number.

        1.  For a given manufacturer, read a speaks file and process it to make it workable.
                - For some stupid reason SPEAKS and CEDNet absolutely MUST export documents as 
                         if they were going to be printed and read on paper.  This probably made sense 
                         10 years ago, but now it just leaves me with everything formatted like a PDF document.
                         For example, an Excel sheet with all of the page headers and footers crammed
                         into clusters of merged cells and it's a nightmare to work with.  It does the
                         same for txt files, but those are much much easier to work with in general, so I go with that.

                -It winds up with a list structured similarly to:

                        CUSTOMER    MFR CAT#    something   something   QUANTITY

                 and these are the only values that i'm interested in.

        2.  Open the current running sales sheet, and for every sale of a matching product 
                 in that list above, put it in the list with all of its information
            -   Each manufacturer's sheet is a little bit different, but contains all
                       of the same information.  Therefore each manufacturer's method is slightly
                       different.  The least similar one is LG, but that only needs to have the 
                       products grouped together and tallied.

        3.  Save the file however it needs to be saved.

        4.  Our vendors used to have us report on our current stock levels of their products.  
                 In the Ironridge and Enphase sections there are swaths of commented-out code that used to do that. 
                 it doesn't do anything useful anymore, but I don't think it hurts to keep around just as a showcase.

'''
from openpyxl import Workbook
from openpyxl import load_workbook
import CEDNetUtils as CED
from datetime import datetime
from datetime import date

from subprocess import check_output
from subprocess import CalledProcessError

from os import system

from time import strftime

path = "C:\\PaulScripts\\"


CUST_ACCT = 0
CUST_NAME = 1
CUST_ADDR = 2
CUST_ZIP = 3
CUST_CITY = 5
CUST_STATE = 4


def do_ironridge(customers, product_numbers, wbname, iridg):
    
    if wbname is None:
        print(("Ironridge POS workbook not found.  Consider editing\n"
                "C:\\PaulScripts\\Solar Reporting\\Workbook_names.txt"))
        return
    pos = run_speaks(customers, path + "Speaks Exports\spksweek.txt", "IRIDG", iridg)
    if pos == -1:
        return -1
        
    print(pos[0])
    
    
    

    iridgwb = load_workbook(path + "Solar Reporting\Weekly\\"+wbname)

    ws = None

    railsheet = iridgwb.get_sheet_by_name("Rail Sales")
    flashsheet = iridgwb.get_sheet_by_name("Flashing Sales")
    correct = False
    for i, each in enumerate(pos):
        cust = []
        for each2 in customers:
            if each2[CUST_ACCT] == each[CUST_ACCT]:
                cust = each2
                break

        if each[2][:2] == 'XR':
            ws = railsheet
            mr=ws.max_row+1
        else:
            ws = flashsheet
            mr = ws.max_row+ 1
        
        '''     
        print(customers[-1])
            
        for i, each2 in enumerate(cust):
            print(str(i) + " " + each2)
        for i, each2 in enumerate(each):
            print(str(i) + " " + each2)
        input("...")'''


        #print(cust)
        #print(each)
        ws.cell(row=mr, column=1).value = "CED Greentech" #Distributor
        ws.cell(row=mr, column=2).value = each[product_numbers[0]]      #Date
        ws.cell(row=mr, column=3).value = each[product_numbers[1]]      #Cat #
        ws.cell(row=mr, column=4).value = each[product_numbers[2]]      #qty
        ws.cell(row=mr, column=5).value = cust[CUST_STATE]      #state
        ws.cell(row=mr, column=6).value = cust[CUST_ZIP]        #zip
        ws.cell(row=mr, column=7).value = "USA"         #country
        ws.cell(row=mr, column=8).value = cust[CUST_NAME]       #Customerf


    #iridgwb.save(path + "Speaks Exports\Ironridge POS {i}.xlsx".format(i=strftime("%B %d %Y")))
    iridgwb.save(path + "Solar Reporting\Weekly\\" + wbname)
    
def do_enphase(customers, product_numbers,wbname, enp):
    if wbname is None:
        print(("Enphase POS workbook not found.  Consider editing\n"
                "C:\\PaulScripts\\Solar Reporting\\Workbook_names.txt"))
        return
    
    pos = run_speaks(customers, path + "Speaks Exports\spksweek.txt", "ENP", enp)
    
    if pos == -1:
        return -1
        
    enpwb = load_workbook(path + "Solar Reporting\Weekly\\"+wbname)
    ws = enpwb.get_sheet_by_name("POS_Data")
    mr = ws.max_row+ 1


    skip = 0

    for i, each in enumerate(pos):
        cust = []

        for each2 in customers:
            if each2[CUST_ACCT] == each[CUST_ACCT]:
                cust = each2
                break
        print(cust)
        #for isi, each2 in enumerate(each):
        #   print(str(isi) + ". " + each2)
        #input("...")
        #quit()

        if int(each[product_numbers[2]]) <= 0:
            skip += 1
            continue
        else:
            ws.cell(row=mr + i- skip, column=4).value =  each[product_numbers[2]]#qty
        
        ws.cell(row=mr + i - skip, column=1).value = "CED Greentech" #Distributor
        ws.cell(row=mr + i- skip, column=2).value = each[product_numbers[0]] #Date
        ws.cell(row=mr + i- skip, column=3).value =  each[product_numbers[1]]#Cat #


        ws.cell(row=mr + i- skip, column=5).value =  cust[CUST_STATE]#State
        ws.cell(row=mr + i- skip, column=6).value =  cust[CUST_ZIP]#Zip
        ws.cell(row=mr + i- skip, column=7).value =  "USA"#Country
        ws.cell(row=mr + i- skip, column=8).value =  each[product_numbers[3]]#Invoice No
        ws.cell(row=mr + i- skip, column=9).value =  "1"#Line #
        ws.cell(row=mr + i- skip, column=11).value =  cust[CUST_NAME]#Name
        ws.cell(row=mr + i- skip, column=12).value =  cust[CUST_ADDR]#Addr
        ws.cell(row=mr + i- skip, column=13).value =  cust[CUST_CITY]#City

        #print(ws.row(mr+i))

    enpwb.save(path + "Solar Reporting\Weekly\\"+wbname)

def do_sma(customers, product_numbers, sma):
    if wbname is None:
        print(("SMA workbook not found.  Consider editing\n"
                "C:\\PaulScripts\\Solar Reporting\\Workbook_names.txt"))
        return
    pos = run_speaks(customers, path+"Speaks Exports\spksmonth.txt", "SMA", sma)
    if pos == -1:
        return -1
        
    smawb = load_workbook(path + "Solar Reporting\Monthly\SMA\SMA Template.xlsx")
    ws = smawb.get_sheet_by_name("POS")
    mr = 9


    for i, each in enumerate(pos):
        #print(each)
        cust = []
        if int(each[-1]) <= 0:
            #skip += 1
            continue
        for each2 in customers:
            if each2[0] == each[0]:
                cust = each2
                break

        print(each)
        print(cust)

        ws.cell(row=mr + i, column=1).value =   each[product_numbers[0]] #Date
        ws.cell(row=mr + i, column=4).value =   cust[CUST_NAME] #Customer
        ws.cell(row=mr + i, column=5).value =   each[product_numbers[3]] #Invoice Num
        ws.cell(row=mr + i, column=6).value =   cust[CUST_ADDR] #Customer Address
        ws.cell(row=mr + i, column=7).value =   each[product_numbers[1]] #Cat #
        ws.cell(row=mr + i, column=8).value =   each[product_numbers[2]] #Quantity


        #print(ws.row(mr+i))

    ws.cell("D5").value = date.isoformat(date.today())

    smawb.save(path + "Solar Reporting\Monthly\SMA\SMA {i}.xlsx".format(i=strftime("%B %Y")))
def do_lg(customers, product_numbers, lg):
    if wbname is None:
        print(("LG POS workbook not found.  Consider editing\n"
                "C:\\PaulScripts\\Solar Reporting\\Workbook_names.txt"))
        return
    
    print(customers[57])
    pos = run_speaks(customers, path+"Speaks Exports\spksmonth.txt", "LG", lg)
    if pos == -1:
        return -1
        
    lgwb = load_workbook(path + "Solar Reporting\Monthly\LG\LG Template.xlsx")
    
    ws = lgwb.get_sheet_by_name("POS Reporting")
    mr = 2

    #Sort POS by Cat Num.

    pos2 = []
    for each in pos:
            #bubble sort AYYYYYYYYYYY
        print(each)
        tf = False
        if int(each[-1]) <= 0:
            #skip += 1
            continue
        for i, each2 in enumerate(pos2):
            if each[ORDER_CAT] <= each2[ORDER_CAT]:
                pos2.insert(i, each)
                tf = True
                break

        if tf:
            continue



        pos2.append(each)



    for each in pos2:
        print(each)


    hrmph = 1
    totalqty = 0
    lastproduct = pos2[0][2].strip()
    extraspaces = 0
    for i, each in enumerate(pos2):
        cust = []

        if each[7] == "0":
            extraspaces -= 1
            continue

        if each[2].strip() != lastproduct:
            #extraspaces += 1

            ws.cell(row=mr+i+extraspaces, column=2).value = "Total:"
            ws.cell(row=mr+i+extraspaces, column=3).value = totalqty

            totalqty = 0
            extraspaces += 2
            lastproduct = each[2].strip()




        for each2 in customers:
            if each2[CUST_ACCT] == each[CUST_ACCT]:
                cust = each2
                break
        
        print(cust)
        ws.cell(row=mr + i + extraspaces, column=1).value = each[product_numbers[0]]
        ws.cell(row=mr + i + extraspaces, column=2).value = each[product_numbers[1]]
        ws.cell(row=mr + i + extraspaces, column=3).value = each[product_numbers[2]]
        ws.cell(row=mr + i + extraspaces, column=4).value = cust[CUST_STATE]
        ws.cell(row=mr + i + extraspaces, column=5).value = cust[CUST_ZIP]
        ws.cell(row=mr + i + extraspaces, column=6).value = each[product_numbers[3]]
        ws.cell(row=mr + i + extraspaces, column=7).value = cust[CUST_CITY]

        totalqty += int(each[7])


        #print(ws.row(mr+i))
    ws.cell(row=mr + len(pos) + extraspaces, column = 2).value = "Total:"
    ws.cell(row=mr + len(pos) + extraspaces, column = 3).value = totalqty


    lgwb.save(path + "Solar Reporting\Monthly\LG\LG {i}.xlsx".format(i=strftime("%B %Y")))
    
def do_solaredge(customers, product_numbers, se):
    if wbname is None:
        print(("SolarEdge POS workbook not found.  Consider editing\n"
                "C:\\PaulScripts\\Solar Reporting\\Workbook_names.txt"))
        return
    pos = run_speaks(customers, path + "Speaks Exports\solaredge.txt", "SE", se)
    if pos == -1:
        return -1
        
    print(pos[0])

    sewb = load_workbook(path + "Solar Reporting\Quarterly\SolarEdge POS.xlsx")




    ws = sewb.get_sheet_by_name("Sheet1")
    mr = ws.max_row+ 1
    fewerspaces = 0

    for i, each in enumerate(pos):
        cust = []
        for each2 in customers:
            if each2[CUST_ACCT] == each[CUST_ACCT]:
                cust = each2
                break



        print(cust)
        print(each)


        if each[-1] == "0":
            fewerspaces -= 1
            continue

        #Customer
        ws.cell(row=mr + i + fewerspaces, column= 1).value = cust[CUST_NAME] #Customer Name
        ws.cell(row=mr + i+ fewerspaces, column=2).value = cust[CUST_ADDR] #address
        ws.cell(row=mr + i+ fewerspaces, column=3).value = each[product_numbers[0]] #date
        ws.cell(row=mr + i+ fewerspaces, column=4).value = each[product_numbers[1]]#cat num
        ws.cell(row=mr + i+ fewerspaces, column=5).value = each[product_numbers[2]]#Qty




    #iridgwb.save(path + "Speaks Exports\Ironridge POS {i}.xlsx".format(i=strftime("%B %d %Y")))
    sewb.save(path + "Solar Reporting\Quarterly\SolarEdge POS.xlsx")
    
def run_speaks(customers, fp, mfr, prod):
    
    while True:
        try:
            f = open(fp)
            break
        except FileNotFoundError:
            print( ("Speaks file not found.  Please run a speaks query for the previous week/month.\n"
                    "Save it either at C:\PaulScripts\Solar Reporting\spksweek.txt, or spksmonth.txt"))
            ch = input("Press enter to continue, or q to return to the previous screen.")
            if ch.lower() == 'q':
                return -1
    print("Gathering speaks data...")       
    lines = []

    for line in f:
        #print(line)
        #input(line[0])
        l = line.split("\t")
        for i in range(0, len(customers)):
            #print(customers[i][0])

            if l[0].strip() == customers[i][0]:
                lines.append(l)
                break
        if l[0].strip() == mfr:
            lines.append(l)


    splits = []
    for each in lines:
        #print(each)
        dd = []
        for each2 in each:
            dd.append(each2.replace('\n', ''))

    #   print(dd)
        splits.append(dd)


    asdf = []
    cs = ""
    for each in splits:
        #print(each)
        #input(each)
        for i in range(0, len(customers)):
            if each[0] == customers[i][0]:
                #print("got new customer")
                #print(each)
                cs = each[0]
                continue
        each.insert(0, cs)
        #input(each)

        asdf.append(each)




    final = []

    for each in asdf:

        if len(each)<= 3:
            continue
        if each[2].strip() not in prod:
            #print("woo")
            continue

        final.append(each)




    #print(final[20])

#   print("Final:" + str(final))

    return final

def ced_stock(fp, mfr, prod):
    f = open(fp)
    final = []
    for line in f:
        #print(line)
    #   input("...")
        ddd = []
        if line[:len(mfr)] == mfr:
            sp = line.split(" ")

            sp2 = sp
            sp = []
            for each in sp2:
                if each == "" or each == "\n":
                    continue


                sp.append(each.replace("\xad", "-"))


            if sp[1] in prod:
                ddd.append(sp[1])
                ddd.append(sp[2])
                try:
                    ddd.append(sp[3])
                    if ddd[2].isnumeric() == False:
                        ddd[2] = 0
                except IndexError:
                    ddd.append(0)


                print(ddd)
                final.append(ddd)

    return final

def get_proper_speaks_numbers(customers, iridg):
    while True:
        product_numbers = []
        f=open("C:\\PaulScripts\\configs\\solar_product_info.txt", 'r')
        
        while len(product_numbers) < 4:

            
            line = f.readline()
            print(line)
            try:
                product_numbers.append(int(line[:2].strip()))
            except ValueError:
                print("nop")
                continue
        

        pos = run_speaks(customers, path + "Speaks Exports\spksweek.txt", "IRIDG", iridg)
        
        
        biggest = []
        
        for each in pos:

            if len(each) > len(biggest):
                biggest = each
        
                
        
        print("0. Date:\t" +        biggest[product_numbers[0]])
        print("1. Cat #:\t" +       biggest[product_numbers[1]])
        print("2. Quantity:\t" +    biggest[product_numbers[2]])
        print("3. Invoice #\t" +    biggest[product_numbers[3]])
        
        correct = input("Is this correct?  y/n")
 
        system('cls')
        
        if correct.lower() == "n":

            lines = []
            lines.append( str(product_numbers[0] ) + "\t\tOrder Date\n")
            lines.append( str(product_numbers[1] ) + "\t\tCat Number\n")
            lines.append( str(product_numbers[2] ) + "\t\tOrder Quantity\n")
            lines.append( str(product_numbers[3] ) + "\t\tInvoice Number\n")

            lines.append("Above are the values that the script reads, and below is an entry from CEDNet's customer file.\n")
            lines.append("Find the number which is wrong on the top list, and replace it with the correct number from the bottom list.\n")
            lines.append("\nSome numbers are supposed to be negative.  It's how I can make Python play nicely with speaks.\n\n")
            lines.append("Do not change the order of the values!\n\n")

            biggest = []
            
            for each in pos:
                if len(each) > len(biggest):
                    biggest = each
            
            
            for i, each in enumerate(biggest):
                if i < 5:
                    
                    lines.append(str(i) +"\t\t"+str(each).strip()+"\n")
                
                else:
                    lines.append(str(i - len(biggest)) +"\t\t"+str(each).strip()+"\n")

                
            f=open("C:\\PaulScripts\\configs\\solar_product_info.txt", 'w')
            
            for each in lines:
                f.write(each)
            
            f.close()
            
            print("I wrote a file file for you to edit to tell me what the values are.  \n"\
                    "I will open it in notepad for you.  The script will continue when you close notepad.")
            
            try:
                check_output(["notepad.exe", "C:\\PaulScripts\\configs\\solar_product_info.txt"]).decode("ascii")
                
            except CalledProcessError:
                print("Error opening notepad.")
                
            system('cls')
            
            continue

        elif correct.lower() == 'y':
            return product_numbers
        else:
            print("invalid choice")
            
def get_workbook_names():
    while True:
        try:
            f = open(path+"Solar Reporting\\workbook_names.txt", 'r')
            break
        except FileNotFoundError:
            print("Workbook reference file not found.  I will create one for you to edit...")
            print("Save and close notepad to continue")
            f=open(path+"Solar Reporting\\workbook_names.txt", 'w')
            
            f.write("#Vendor: File Name\n")
            f.write("Ironridge:\n")
            f.write("Enphase:\n")

            f.close()
                
            try:
                check_output(["notepad.exe", "C:\\PaulScripts\\Solar Reporting\\workbook_names.txt"]).decode("ascii")
                
            except CalledProcessError:
                print("Error opening notepad.")
    
    fns = [None, None]
    for line in f:
        if line[0]=='#':
            continue
        sp = line.strip().split(":")
        if sp[0] == "Ironridge":
            fns[0] = sp[1].strip()
        if sp[0] == "Enphase":
            fns[1] = sp[1].strip()

    
    return fns
            
def get_products():
    iridg = []
    enp = []
    sma = []
    lg = []
    se = []
    
    while True:
        

        try:
            f = open("C:\\PaulScripts\\Solar Reporting\\products.txt", 'r')
            break
        except FileNotFoundError:
            f = open("C:\\PaulScripts\\Solar Reporting\\products.txt", 'w')
            f.write("MFR\tCATNUM\n")
            f.close()
            print("Solar product file not found.  Creating one for you to edit.")
            print("Add any products you would like to see to this file in notepad.\nClose notepad to continue.")
            try:
                check_output(["notepad.exe", "C:\\PaulScripts\\Solar Reporting\\products.txt"]).decode("ascii")
                
            except CalledProcessError:
                print("Error opening notepad.")
                
    for line in f:
        
        sp = line.strip().split()
        try:
            if sp[0] == 'IRIDG':
                iridg.append(sp[1])
            if sp[0] == 'ENP':
                enp.append(sp[1])
            if sp[0] == 'SMA':
                sma.append(sp[1])
            if sp[0] == 'SE':
                se.append(sp[1])
            if sp[0] == 'LG':
                lg.append(sp[1])
        except IndexError:
            pass
            
    return iridg, enp, sma, se, lg
        


def run():
    
    iridg, enp, sma, se, lg = get_products()
    
    print("\n\nSolar Reporting\n")
    #csdat = get_customers()
    #if csdat == False:
    #   return
    customers = CED.get_customers()
    #print(customers[1])
    product_numbers = get_proper_speaks_numbers(customers, iridg)
    
    wbnames = get_workbook_names()

    ch = None
    while True:
        opt = input("Weekly? Y/N").upper()
        if opt == "Y":
            ch = do_ironridge(customers,product_numbers, wbnames[0], iridg)
            if ch == -1:
                return
            ch = do_enphase(customers,product_numbers,wbnames[1], enp)
            
            break
        elif opt == "N":
            break

        print("invalid option\n\n")

    while True:
        opt = input("Monthly? Y/N").upper()
        if opt == "Y":
            ch = do_sma(customers,product_numbers, sma)
            if ch == -1:
                return
            ch =do_lg(customers,product_numbers, lg)
            break
        elif opt == "N":
            break

    while True:
        opt = input("Quarterly? Y/N").upper()
        if opt == "Y":
            ch = do_solaredge(customers,product_numbers, se)
            if ch == -1:
                return
            break
        elif opt == "N":
            break
            
            
if __name__ == "__main__":
    run()
else:
    print("imported solar.py")
 