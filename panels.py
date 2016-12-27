#panelstock

from openpyxl import Workbook
from openpyxl.worksheet import dimensions

import datetime
import CEDNetUtils as CED

mfrs = ["LG", "HYU", "SWLD", "QCELL", "JINKO"]
def main():
	
	lns = get_lines()
	print(lns)
	write_book(lns)
	
	
def get_lines():

	prds = CED.get_products()

	pgs = []

	for each in prds:
		if each[0] in mfrs:
			pgs.append(each)
	
	return pgs
					
def write_book(lns):
	wb = Workbook()
	ws = wb.active
	ws.column_dimensions['A'].width = 7
	ws.column_dimensions['B'].width = 23
	ws.column_dimensions['C'].width = 40
	ws.column_dimensions['D'].width = 10
	ws.column_dimensions['E'].width = 10
	
	ws.append(["MFR", "CAT #", "DESC", "OH QTY", "AVAIL."])
	
	for each in lns:
		ws.append(each)
		
	ws.cell("F1").value = "Updated:"
	ws.cell("G1").value = datetime.date.today()
	wb.save("C:\\Users\\pgallagherjr\\Dropbox\\Panel Stock\panels.xlsx")
	
	
	
main()
