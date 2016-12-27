#cyclecounting.py
'''
This script is for keeping track of stock quantities.
It works as follows:
	1.	Loads the cycle counting workbook, and compares it against the CEDNet
		data files, and makes sure that any new products have been added to the list.
	2.	Finds each row that hasn't been checked yet, and returns a list of those row indeces
	3.	Randomly chooses at least 25 of the indeces to check, returns a list of integers.
		- It also checks to see if our our system has a positive quantity of the chosen product onhand
		- If the product quantity is zero, increment a counter to keep the number of products with stock onhand at 25.
		- ie, one product with zero onhand stock will result in 26 entries being chosen.  2 -> 27.  etc.
	4.	Sorts the returned list of integers, then grabs the row data.
		Adds a date to column 6 to show it has been checked.
	5.	Spits out an excel spreadsheet of the chosen products for me to check.
'''


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet import dimensions
from openpyxl.styles import Side, Border

import CEDNetUtils as CED

from datetime import datetime
from datetime import date
from time import strftime

import math
import random

#List of mfrs with things that are very hard to count so it won't return those.
exclude = ["FLEX", "LIQ", "COND", "WIRE", "HYU", "LG", "SWLD", "CORD"]



def do_product_book(wb):
	'''
	opens the cycle workbook, makes sure any new product in SPKPRDDT.lsq are added to the list
	'''

	try:
	
		f = open("C:\\Invsys\\Algorithm\\SPKPRDDT.lsq")

	except FileNotFoundError:
	
		print("CEDNet data files not found.  \nPlease run the Data Files export function in CEDNet")
		print("It can be found at Maintainance > Product > Data Files")
		return
		
	prd = CED.get_products()



	ws = wb.active
	print(ws.max_row)
	#for every product in SPKPRDDT.lsq, make sure it is in the list.
	index = 0
	for i, p in enumerate(prd):

		found = False


		#Since any new products are just appended to the end of the list, it's not in alphabetical order,
		#it needs to check to the very end of the list to be sure if its there or not
		#I'll probably make the workbook sort itself sometime soon.
		
		for j in range(1, ws.max_row+1):

			comp = []
			try:
				comp = [ws.cell(row=j, column = 1).value.strip(), str(ws.cell(row=j, column=2).value).strip()]
			except AttributeError:
				pass
				continue
			if comp == []:
				break


			if  comp == [p[0].strip(), p[1].strip()]:

				found = True

				#need to update the onhand quantity to most recent
				ws.cell(row=j,  column = 4).value = p[3]
				#update the bin location too
				ws.cell(row=j, column=5).value = p[4]

				break

		#if it's not found, append it to the list.
		if found == False:
			index += 1

			p2=[]
			p2.append(p[0])
			p2.append(p[1])
			p2.append(p[2])
			p2.append(p[3])
			p2.append(p[4])

			
			ws.append(p2)

	print(ws.max_row)

	#ws = sort_wb(wb)

	return ws

def get_items_to_check(ws):

	items = []

	#get the list of row indeces that haven't been checked.
	pool = get_pool(ws)

	#get at least 25 random numbers.
	zeroes = 0
	rands = []
	while len(rands) - zeroes <=	40:
		rnd = random.randint(0,len(pool))

		#if the system says we have none of the chosen product on hand, increment a counter to offset the number of zeroes.
		if ws.cell(row=rnd, column = 4).value == "0":
			zeroes+=1
		rands.append(rnd)

	rands.sort()
	print(zeroes)
	print(len(rands))
	print(rands)


	#	for each random number chosen, take the corresponding value from pool[],
	#	and then take THAT corresponding row from the workbook.
	for each in rands:
		i = []
		try:
			p = pool.pop(each)
		except IndexError:
			continue

		i.append(ws.cell(row=p, column = 1).value ) #mfr
		i.append(ws.cell(row=p, column = 2).value )	#cat num
		i.append(ws.cell(row=p, column = 3).value )	#desc
		i.append(ws.cell(row=p, column = 5).value )	#bin
		i.append(ws.cell(row=p, column = 4).value ) 	#qty
		i.append(" ")


		#note that this row has been counted
		ws.cell(row=p, column=6).value = date.isoformat(date.today())

		items.append(i)
		print(str(len(items)) + " " + str(i))

		if len(pool) == 0:
			break

	#return items
	return items

def get_pool(ws):
	'''
	Check each row for a date in column 6, and build a list of rows without it.
		- Doesn't include rows containing products in Exclude[]

	If every row has been checked, reset the columns and we start from the beginning.

	Todo: Maybe it should find the earliest date and return those.  Check what hasn't been checked in the longest time.

	'''
	pool = []
	while True:
		i=0
		print("building pool")
		for i in range(1, ws.max_row):
			try:
				if ws.cell(row=i, column = 1).value.strip() in exclude:
					continue
			except AttributeError:
				continue
			if ws.cell(row=i, column = 6).value == None:
				pool.append(i)
		if len(pool) != 0:
			print("pool ok")
			return pool

		print("pool is empty - resetting")
		for i in range(1, ws.max_row):

			ws.cell(row=i, column = 6).value = None

def save_as_excel(s):


	wb = Workbook()
	ws = wb.active

	ws.column_dimensions['A'].width = 7
	ws.column_dimensions['B'].width = 23
	ws.column_dimensions['C'].width = 40
	ws.column_dimensions['D'].width = 10
	ws.column_dimensions['E'].width = 10


	header = (["MFR", "CAT #","DESCRIPTION","BIN", "#","ACTUAL" ])
	side = Side(border_style='thin', color="00000000")
	cellborder = Border(left=side, right=side, bottom=side, top=side)

	ws = CED.write_excel_sheet(header = header, list = s, sheet=ws, border=cellborder)

	wb.save("C:\\PaulScripts\\cycle_out.xlsx")

def sort_wb(wb):

	ws = wb.active
	mr = ws.max_row
	mc = ws.max_column
	rows = []

	for i in range(1, mr):
		r=[]
		r.append(ws.cell(row=i, column = 1).value )
		r.append(ws.cell(row=i, column = 2).value )
		r.append(ws.cell(row=i, column = 3).value )
		r.append(ws.cell(row=i, column = 5).value )
		r.append(ws.cell(row=i, column = 4).value )
		rows.append(r)

	newrows = []
	for each in rows:
		for neweach in newrows:
			if each[0].strip() > neweach[0].strip():
				continue

			if each[1].strip() > neweach[1].strip():
				continue

			if each[1].strip() == neweach[1].strip():
				break

			newrows.append(each)
			break


	nws = wb.create_sheet()

	for each in newrows:
		nws.append(each)


	return nws

def run(skip_prompt = False):

	wb = None
	try:

		wb = load_workbook("C:\\PaulScripts\\Inventory Checking\\cycle.xlsx")

	except FileNotFoundError:
		wb=Workbook()


	#Prepares the workbook for processing.
	ws = do_product_book(wb)
	#wb.save("C:\\PaulScripts\\Inventory Checking\\cycle.xlsx")

	if skip_prompt != True:
		ch = input("Would you like to get today's cycle? Y/n...")
		if ch.lower() == "n":

			print("Skipping cycle.")


		else:
			#Retrieves a list of at least 25 items to go count.
			spit = get_items_to_check(ws)
			#Output the list
			save_as_excel(spit)		
	else:
		spit = get_items_to_check(ws)
		#Output the list
		save_as_excel(spit)
	wb.save("C:\\PaulScripts\\Inventory Checking\\cycle.xlsx")

if __name__ == "__main__":
	run(True)
else:
	print("imported cycle.py")
