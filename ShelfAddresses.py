#ShelfAddresses.py
'''
    My greatest *working* achievement of my time at CED

    In order to maintain a warehouse full of products, you need to know where on the shelf everything is.
    I was tasked with doing that for all of the products in our entire warehouse.  Around 3,300 products.
    CEDNet has no way to import a list of shelf locations.  They must all be input manually.

    Bump.
    That.

    I wrote this program to automate the process of entering bin locations.  Trying to do it all manually
    would have taken me more than twice as long as it took to get this working properly.

    This code is heavily commented already because i personally think it's rather obtuse.  Here is a summary...

    The process is as follows:

    0.    In order for this to work, the CEDNet window must be oriented in a very specific way.
        This is because the program creates mouse and keyboard inputs, and if the buttons are in the wrong spot
        it will just run without actually doing anything.

    1.    Open the Stock Status spreadsheet with updated bin locations in it

    2.    Read through the rows until it finds a cell with an updated location.
        Some products that don't have set shelf locations such as wire, conduit, and solar panels don't need it.  They are skipped.

    3.    Prepares CEDNet to receive the location
        -    This means clicking on and entering the manufacturer in the correct box, pressing the tab key until the typing cursor
            is in the bin location field, and typing it in.

    4.    Pressing enter through the products until it arrives at the item to enter.
        -    CEDNet is silly and won't immediately type what you press while the field is highlighted.  Instead, the first keyboard
            input activates the field and everything after that is what is entered.  To get around this, the script just types a few
            arbitrary letters, and then backspaces them before typing in the real value.

    5.  Continues scrolling through the products in case there are more products under this manufacturer

    6.  At the end of the mfr, press the save button, wait for the saving to complete, and then press OKAY and continue.
        -    I did some math and found out that CEDNet and our server saves bin locations at a rate of twenty per second.
        -    To be safe, the script waits for .075 seconds times the number of products it just scrolled through.


    There is the occasional bug that I cant track down, and I don't know what I could do about it.
    Every once in a while the computer will hiccup and an input or two will be lost.    This throws off the entire procedure.

    Every time the script completes a manfacturer, it takes note of the row it worked up until.  This way if I have to stop
    in the middle, it will pick up where it left off.


'''
import os

import time

import win32api
import win32con

import math

import win32
import win32com.client
import ctypes


from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime



path = "C:\\PaulScripts\\"

wb = Workbook()
bookpath = 'C:\\PaulScripts\\shelf locations.xlsx'

orders = []

tsleep = .2 #.01 for fast, 1+ for debugging
shell = None

def type_bin_locations():

    # =IF(AND(G1<>"", H1<>""), UPPER(CONCATENATE(G1, ".", H1, IF(I1<>"", CONCATENATE(".", I1),""))),"")


    #Reads strings out of stock status.xlsx and types them out into CED Net.
    #Uses windows API calls to click around the screen and type.

    #backup_stock_workbook()
    wb = load_workbook("C:\\PaulScripts\\This Week's Stock Status.xlsx", data_only=True)
    
    #make sure the workbook is closed before beginning.
    while True:
        try:
            wb.save("C:\\PaulScripts\\This Week's Stock Status.xlsx")
            break
        except:
            input("Please close the stock status excel document.  Press enter to continue")
    
    ws = wb.get_sheet_by_name("Sheet")

    print("got workbook")


    #index of which line in the spreadsheet the program has already ran until.
    DONE_ALREADY = get_done_already(ws)
    
    #click where the taskbar should be  to bring focus to CEDNet
    click(233, 14)

    mfr = ""
    items = 0

    mfr_i_just_did = "-start-"


    for i in range(DONE_ALREADY, ws.max_row):
        items = 0
        #bin = ""

        if do_i_skip_this_item(ws, i, mfr_i_just_did) == True:
            continue

        #if it's gotten this far, there is a new manufacturer with items that have a new bin location

        print("****************************************")
        print("Found item with new bin location" +str(i) + ": " + ws.cell(row=i, column=1).value + " " + str(ws.cell(row=i, column=2).value))
        print("****************************************")


        #take note of the manufacturer that we just ran through so we don't do it again after this
        mfr_i_just_did = ws.cell(row=i, column=1).value.strip()

        #new manufacturer to run through
        mfr = ws.cell(row=i, column = 1).value.strip()

        #find the first line of the current manufacturer
        mfrstart = get_mfr_start(ws, i)

        #find the last line of the current manufacturer
        mfrend = get_mfr_end(ws, i)

        #total number of items in for this manufacturer.  Only used to help time the sleep after pressing save.
        items = get_mfr_items(ws, i)
        
        print("Items: " +str(items))



        print(str(items) + " for this manufacturer.")

        setup_cednet_for_new_manufacturer(mfr)

        write_in_cednet(ws, mfrstart, mfrend)

        print("done with this mfr")
        
        
        # ########
        saveStockStatus(mfrend, wb, ws, items)
        # ########

    #Save the workbook to start at row 0 next time, because we finished the whole thing.
    saveStockStatus(0, wb, ws, 0)

def saveStockStatus(i, wb, ws, items):
    
    

    click(40, 70)    #click save button in CEDNet


    #Set cell value to checkpoint where we've gotten to so far.
    ws.cell("m1").value = i
    while True:
        try:
            print("SAVING - DONT QUIT YET")
            wb.save("C:\\PaulScripts\\This Week's Stock Status.xlsx")
            print("DONE SAVING")
            
            break
        except:
            input("Please close the stock status excel document.  Press enter to continue")


    time.sleep(items * .075)   #during normal operations it takes CEDNet around .05 seconds per entry to save a matrix.
                                        #Wait a little bit longer just in case.


    click(int(ctypes.windll.user32.GetSystemMetrics(0)/2),
          int(ctypes.windll.user32.GetSystemMetrics(1)/2) + 100)                #Click the OK button on the dialog button that appears.
    items = 1

def write_item(ws, i, item, loc):

    ws.cell(row=i, column=2).value = item
    ws.cell(row=i, column=3).value = loc

def loadBook():
    return load_workbook(bookpath, data_only = True, read_only = False)

def pause():
    #because typing input("askdjfhlaskjdfh") every time i need to pause is stupid
    input("...")

def click(x, y):

    #Call the windows API to click the screen at coordinate (X, Y)

    print("clicked: {a}, {b}".format(a=x, b=y))
    win32api.SetCursorPos((x, y))                                                            #set Cursor Position
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x, y, 0, 0)    #press mouse button down
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x, y, 0, 0)            #release mouse button
    time.sleep(tsleep*4)        #Sleep for four * sleepcoefficient

def mwheel(x):
    #scroll the mouse wheel upward for X number of clicks
    print("mousewheeling for 120 {x} times".format(x=x))
    for i in range(0, x):

        win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL, 0, 0, 120, 0)

    time.sleep(tsleep*4)

def press_keys(bin):
    #print("Typing: " + bin)
    time.sleep(tsleep)
    win32com.client.Dispatch("WScript.Shell").SendKeys(bin)

def type_out(shelf_bin):
    #cheap hack to get around CEDNet being stupid
    #types 2 arbitraty letters into the box and then erases them to ensure that the 
    #bin location box is in focus before writing it
    #If this doesn't happen the first character of the bin location gets 
    #truncated and is a major pain

    time.sleep(tsleep)
    press_keys("Ff")
    time.sleep(tsleep)
    press_keys("{BACKSPACE}")
    time.sleep(tsleep)
    press_keys("{BACKSPACE}")

    press_keys(shelf_bin)

def backup_stock_workbook():
    #Janky way to back up the file.  It's so large that sometimes \
    #I might terminate the program in the middle of saving so it's 
    #nice to have a backup just in case.

    wb2 = Workbook()

    try:
        wb2= load_workbook(path + "Shelving Addresses\\stock status.xlsx")
        wb2.save(path + "Shelving Addresses\\stock status backup.xlsx")

        print("got it")
    except KeyError:
        print("woops")
        wb2= load_workbook(path + "Shelving Addresses\\stock status backup.xlsx")
        wb2.save(path + "Shelving Addresses\\stock status.xlsx")

def get_done_already(ws):
    #returns the last manufacturer that was ran through so it can pick up where it left off in case of interruption
    
    da = ws.cell("m1").value
    

    
    if da is None or da == 0: 
        da = 2
    print("Starting with Mfr: "+ ws.cell(row=da, column=1).value)
    
    

    return da

def get_mfr_start(ws, i):
    ms = 1
    mfr = ws.cell(row=i, column=1).value.strip()
    #find the first case of the mfr in the list.
    for j in range(i, 1, -1):
        if ws.cell(row=j, column = 1).value.strip() == mfr:
            continue

        ms = j+1
        break
        

    if ms == 0:
        ms = 1

    return ms

def get_mfr_items(ws, i):
    me = i
    mfr = ws.cell(row=i, column=1).value.strip()
    for j in range(i, ws.max_row):
        me=j - i - 1

        if ws.cell(row=j, column = 1).value.strip() != mfr:

            break


    return me

def get_mfr_end(ws, i):
    mfr = ws.cell(row=i, column=1).value.strip()
    for j in range(i, ws.max_row):

            
        if ws.cell(row=j, column = 1).value.strip() != mfr:

            break
            
        if ws.cell(row=j, column = 6).value != None:
            last = j
    return last
def setup_cednet_for_new_manufacturer(mfr):
    #manipulate cednet to search for a new manufacturer and prepare for entering new bin locations

    click(177, 211)    #click the Manufacturer text box in CEDNet
    for a in range(0, 10):
        #erase whatever is already in the box.  Manufacturers can have only up to 5 letters
        #so 5 backspaces and 5 deletes should be sufficient to clear whatever is in that box

        press_keys("{BACKSPACE}")
        press_keys("{DELETE}")

    press_keys(mfr.strip() + "{F10}") #F10 is the search button in CEDNet

    press_keys("{TAB}")

    press_keys("{TAB}")

    press_keys("{TAB}")

    press_keys("{TAB}")

    press_keys("{TAB}")

def do_i_skip_this_item(ws, i, mfr_i_just_did):
    #if mfr is the same manufacturer as the one I just did, skip to the next row
    try:
        if ws.cell(row=i, column=1).value.strip() == mfr_i_just_did:
            return True

        #if cell an item does not have a new bin location, skip to the next row
        if ws.cell(row=i, column=6).value is None or ws.cell(row=i, column=6).value == "":
            return True
    except AttributeError:
        return True


    return False

def write_in_cednet(ws, mfrstart, mfrend):
    #for each item within the mfr range, write it in CEDnet

    for j in range(mfrstart, mfrend+1):

        bin = ws.cell(row=j, column=6).value

        #if there is no bin location for an item, skip that item.
        if bin is None or bin == "":
            press_keys("{ENTER}")
            continue
    #    print(ws.cell(row=j, column=2).value + " " + str(bin))

        print("{i}. {name} - {bin}".format(i=j, name= ws.cell(row=j, column=2).value, bin=bin))
        type_out(bin)
        press_keys("{ENTER}") #continue to the next row

        ws.cell(row=j, column = 5).value = bin
        ws.cell(row=j, column = 6).value = None

def done():
    for i in range(0, 25):
        print("\n")

    print("done!")

    for i in range(0, 25):
        print("\n")

def run():

    print("\n\n\n\n")
    print("open CEDNet to the matrix mass maintainance screen")
    print("position the screen so that CEDNet takes up the left half of the screen")
    print("click on the title bar, and drag it to the left of the screen and it should size itself automatically.\n")
    input("\npress enter when you are ready")

    type_bin_locations()

    done()
    os.system("cls")
if __name__ == "__main__":
    run()
else:
    print("imported ShelfAddresses.py")