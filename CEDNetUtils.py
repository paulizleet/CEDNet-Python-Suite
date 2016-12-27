#CEDNetUtils.py
#Usually I copy/paste code from one script to the next.
#I'll put all common code in here instead.

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet import dimensions
from openpyxl.styles import Side, Border
from subprocess import check_output
from subprocess import CalledProcessError

from os import system

from datetime import datetime
import math
import random

def split_product_lines():

		
	
	f = get_ced_file("SPKPRDDT.lsq")
	
	prd = []
	splits = []
	for line in f:
		splits.append(line.split("|"))
	return splits

def get_mfrs():
	mfr = []
	
	f = get_ced_file("SPKMFRDT.lsq")

	for line in f:
		mfr.append(line.split("|"))
	return mfr
	
	
def get_products():

	prd = []
	f=get_ced_file("SPKPRDDT.lsq")
			
	prd = []
	splits = []
	for line in f:
		splits.append(line.split("|"))
	for line in splits:
		p = []
		
		p.append(line[1].strip()) #mfr
		p.append(line[2].strip()) #cat num
		p.append(line[5].strip()) #Description
		p.append(line[15].strip()) #onhand qty
		p.append(line[29].strip()) #bin location
		prd.append(p)
		
	return prd
	
def get_customers():
	while True:
		cs=get_ced_file("SPKCUSDT.lsq")
		
		
		
		customer_numbers = []
		
		f=open("C:\\PaulScripts\\configs\\customer_info.txt", 'r')
		
		while len(customer_numbers) < 7:

			line= f.readline()

			try:
				customer_numbers.append(int(line[:1].strip()))
			except ValueError:
				continue
				

			

		customers = []
		cline = []
		for line in cs:
			cline = []
			#print(line)
			l = line.split("|")
			cline.append(l[customer_numbers[0]]) #ACCT
		#	print("|" + l[customer_numbers[1]] + "|")
			cline.append(l[customer_numbers[1]]) #NAME
			cline.append((l[customer_numbers[2]]+ " " + l[customer_numbers[3]]).strip()) #Addr
			cline.append(l[customer_numbers[4]]) #city
			cline.append(l[customer_numbers[5]]) #zip
			cline.append(l[customer_numbers[6]]) #State


			customers.append(cline)

		correct = False

		print("0. Account Num: "        +   customers[34][0])
		print("1. Customer Name: " 		+   customers[34][1])
		print("2. Customer Addr:"       +   customers[34][2])
		print("3. Customer City:"       +   customers[34][5])
		print("4. Customer State: " 	+   customers[34][4])
		print("5. Customer Zip: "       +   customers[34][3])

		correct = input("Is this correct?  y/n")

		if correct.lower() == "n":

			lines = []
			lines.append( str(customer_numbers[0] ) + "\t\tAccount Number\n")
			lines.append( str(customer_numbers[1] ) + "\t\tCustomer Name\n")
			lines.append( str(customer_numbers[2] ) + "\t\tAddress Line 1\n")
			lines.append( str(customer_numbers[3] ) + "\t\tAddress Line 2\n")
			lines.append( str(customer_numbers[4] ) + "\t\tCustomer city\n")
			lines.append( str(customer_numbers[5] ) + "\t\tCustomer zip\n")
			lines.append( str(customer_numbers[6] ) + "\t\tCustomer state\n\n\n")
			
			lines.append("Above are the values that the script reads, and below is an entry from CEDNet's customer file.\n")
			lines.append("Find the number which is wrong on the top list, and replace it with the correct number from the bottom list.\n")
			lines.append("Do not change the order of the values!\n\n")

			
			f=get_ced_file("SPKCUSDT.lsq")
			
			csline= []
			for i in range(0, 150):
				csline=f.readline()

			
			for i, each in enumerate(csline.split("|")):
				lines.append(str(i) +"\t\t"+str(each).strip()+"\n")
				
			f=open("C:\\PaulScripts\\configs\\customer_info.txt", 'w')
			
			for each in lines:
				f.write(each)
			
			f.close()
			
			print("I wrote a file file for you to edit to tell me what the values are.  \n"\
					"I will open it in notepad for you.  The script will continue when you close notepad.")
			
			try:
				check_output(["notepad.exe", "C:\\PaulScripts\\configs\\customer_info.txt"]).decode("ascii")
			except CalledProcessError:
				print("Error opening notepad.")
				
			system('cls')
				
			
			continue

		elif correct.lower() == 'y':
			return customers
		else:
			print("invalid choice")
			
			
def get_ced_file(filename):
	while True:
		try:    
			
			return open("C:\\Invsys\\Algorithm\\" + filename)
			
		except FileNotFoundError:
			print("CEDNet data files not found.  \nPlease run the Data Files export function in CEDNet")
			print("It can be found at Maintainance > Product > Data Files")
			input("Press enter when ready...")
			continue
	
def write_excel_sheet(header = None, list = None, sheet=None, border=None):
	
	offset = 1
	if header != None:
		sheet.append(header)
		offset = 2
		
	for i, each in enumerate(list):
		sheet.append(each)
		if border != None:
			for j in range(1, sheet.max_column):
				sheet.cell(row=i+offset, column=j).border = border
	return sheet

		
	

	