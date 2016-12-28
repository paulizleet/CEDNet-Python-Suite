'''


	PricingMatrix.py
	
	
	Reads in a given pricing matrix, and alters the prices based on current vendor pricing


	Surprisingly, CEDNet has a method of importing and exporting customer pricing matrices.
	However, it doesn't have a way of linking customer pricing with vendor cut sheets.

	The price we buy our wire at is dependent on the price of copper, which fluctuates with the market.
	As such, we need to change the sale price of much of our wire stock every few days.

	It was also used to help create a massive pricing matrix to give a default price for every product,
	just in case it wasn't in a matrix for any customer.

	This was written to solve that problem.

	CEDNet spits the data out into a godawful flat file, and only accepts input from a properly formatted one.
	It does contain the data we need though.

	The script goes like so:

	0.	Before the program may be run, several things must happen.
		-	CEDnet must have exported the requested matrix as a template to read from.
		-	An excel spreadsheet of where the script can find the latest price for a given product must be created manually.
			This is discussed later.
	1.	Read the template matrix to gather a list of products
	2.	We sell our wire at two price points for different customers: 10% for preferred customers, and 15% for everybody else
		For those vendors, churn out a matrix
	3.	These two tiers of customers each have their own matrix, so any products must be fitted to the new matrix.
	4.


'''

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple
from openpyxl.formula import Tokenizer

from decimal import *

import os
import math

#import basematrix
TOP_CUSTOMER = .9
REG_CUSTOMER = .85

WIRE_VENDOR = 1
PIPE_VENDOR = 2
path = "C:\\PaulScripts\\Wire Matrix\\"




def do_matrix(vendor):
		
	if vendor == WIRE_VENDOR:
		matrix_in = "wiret_BACKUP.lsq"
		destination = "WWIRE"
		top_dest  = "WWIRT"
		sheet = "WIRE"
	elif vendor == PIPE_VENDOR:
		matrix_in = "pipe.lsq"
		destination = "PIPE"
		top_dest  = "PIPE1"
		sheet= "PIPE"
	else:
		print("Invalid vendor: " + str(vendor))
		return
	
	margin = .85
	writedest = destination
	runs = 0
	#top vendor
	while True:
	
			
		productList = []
		
		while True:
			try:
				f = open(path + matrix_in, 'r')
				break
			except FileNotFoundError:
				print("Please export the matrix you would like to update to C:\\PaulScripts\\Wire Matrix\\{m}".format(m=matrix_in)
				ch=input("Or type q to quit.")
				if ch.lower() == "q":
					return -1
					
				
				
		
		
		productList = readMatrix(f)
		
		f.close()

		productList = cleanMatrix(productList)
		productList = changeDestinationMatrix(writedest, productList)
		productList = updatePrices(productList, sheet, vendor, margin)

		f = open("C:\\PaulScripts\\Wire Matrix\\IMPORT_MATRIX.lsq", 'a')
		writeMatrix(f, productList)
		f.close()
		runs+=1
		if runs == 1:
			margin = .9
			writedest = top_dest
			continue
		break
			
def readMatrix(f):

	pl = []

	for line in f:
		pl.append(parseLine(line))


	return pl

def writeMatrix(f, productList):

	for line in productList:
		wline = ""
		wline+=line[0]
		wline+=line[1]
		wline+=formatPrice(line[2])
		wline+=line[3]
		wline+="\n"
		f.write(wline)

def parseLine(line):

		pl = []
		pl.append(line[:11])
		pl.append(line[11:33])
		pl.append(float(line[33:48]) /100)
		pl.append(line[48:97])

		return pl

def formatPrice(price):
	zeroes = "000000000000000"
	if price is None:
		return zeroes

	price=int((float(price)*100))
	price = str(price)
	price = zeroes[:-len(price)]+price
	return price

def updatePrices(pl, SHEET = None, VENDOR = None, MARGIN = None):

	refsheet = None
	if VENDOR == WIRE_VENDOR:
		vendorBook = loadSouthwireBook()
		refbook = loadWireReferenceBook()
		refsheet = refbook.get_sheet_by_name(SHEET)
		hotbook = Workbook()
		hs = hotbook.create_sheet("Wire Pricing")
		hs.append(["Type", "Price per 100ft"])

	if VENDOR == PIPE_VENDOR:
		vendorBook = loadConduitWorkbook()
		refbook = loadWireReferenceBook()
		print(vendorBook.get_sheet_names())
		refsheet = refbook.get_sheet_by_name(SHEET)
		hotbook = Workbook()
		hs = hotbook.create_sheet("Pipe Pricing")


	j = 1
	for i, each in enumerate(pl):

		if getCellValueString(refsheet, r=j, c=3) != "DNU":

			prc = getPrice(each[1], refsheet, vendorBook, j)

			if prc == -1:
				each[2] = 0.0
			else:
				try:
					each[2] = round(prc / MARGIN, 2)
				except TypeError:
					print(each)
					print("There was an error trying to update this price.  \n"\
					"Probably because the reference book entry pointed to an empty cell.\n\n"\
					"Double check that the new sheets you were given follow the same format,\n"\
					"and make sure that all of the references point to the right place.")
					
					print("I'm skipping this entry for now, the price will remain the same. \n"\
					"I'm also adding this to a log of errors for you to look at afterwards.")
					
					f=open("C:\PaulScripts\Wire Matrix\wire_matrix_error_log.txt", 'a')
					errorline = "Reference Book \t"
					errorline += each[1]
					errorline += "\t Coordinate error\n"
					f.write(errorline)
					f.close()
					
					continue

				hs.append([each[1], prc])
		else:
			if MARGIN == TOP_CUSTOMER:
				each[2] = each[2]*.85
				each[2] = each[2]/.9


		j+=1

	hotbook.save('C:\\PaulScripts\\Wire Matrix\\hotbook.xlsx')

	return pl

def getPrice(product, rb, vb, i):

	for i in range(1, rb.max_row + 1):

		if product.strip() == getCellValueString(rb,r=i,c=1).strip():

			c = getCoordinates(i, rb)


			if c[0] == "DNU":
				return -1.0
			print(c)
			print(getCellValueString(rb, "A{0}".format(i) )+ " " + str(c) + " " + str(getVendorCost(c, vb)) + "\n")

			return getVendorCost(c, vb)


	return -1.0

def getCoordinates(i, rb):

	return (getCellValueString(rb, "C{0}".format(i)), getCellValueString(rb, "D{0}".format(i)) + getCellValueString(rb, 'E{0}'.format(i)))

def getVendorCost(c, vb):



	if c[0] == "CU":
		return vb[0].cell(c[1]).value

	if c[0] == "CUspc":
		return vb[1].cell(c[1]).value

	if c[0] == "Bare":
		return vb[2].cell(c[1]).value

	if c[0] == "AL":
		return vb[3].cell(c[1]).value

	if c[0] == "AlSpc":
		return vb[4].cell(c[1]).value

	if c[0] == "MC":
		print(vb[5].cell(c[1]).value)
		return vb[5].cell(c[1]).value

	return vb.get_sheet_by_name(c[0]).cell(c[1]).value

def loadPipeReferenceBook():
	return load_workbook('C:\\PaulScripts\\Wire Matrix\\referenceBook.xlsx', data_only = True)

def loadConduitWorkbook():
	return load_workbook('C:\\PaulScripts\\Wire Matrix\\conduit.xlsx', data_only = True)

def loadWireReferenceBook():
	return load_workbook('C:\\PaulScripts\\Wire Matrix\\referenceBook.xlsx', data_only = True)

def loadSouthwireBook():
	wirebook = load_workbook('C:\\PaulScripts\\Wire Matrix\\wirebooks\\sw.xlsx', data_only = True)
	mc = load_workbook('C:\\PaulScripts\\Wire Matrix\\wirebooks\\mc.xlsx', data_only = True)
	sheets = [None]*6
	for each in wirebook.get_sheet_names():
		if "commercial cu" in each.lower():
			sheets[0] = wirebook.get_sheet_by_name(each)
		if "residential cu" in each.lower():
			sheets[1] = wirebook.get_sheet_by_name(each)
		if "bare copper" in each.lower():
			sheets[2] = wirebook.get_sheet_by_name(each)
		if "commercial al" in each.lower():
			sheets[3] = wirebook.get_sheet_by_name(each)
		if "residential al" in each.lower():
			sheets[4] = wirebook.get_sheet_by_name(each)
	ar = mc.get_sheet_names()

	for each in mc.get_sheet_names():
		if "mc alum 14-10 awg" in each.lower():
			sheets[5] = mc.get_sheet_by_name(each)
			break

	print(sheets)
	return sheets

def getCellValueString(worksheet,coor = None,  r=None, c=None):
	if coor is None:
		if (r is None or c is None):
			msg = "You have to provide a value either for " \
					"'coordinate' or for 'row' *and* 'column'"
		return str(worksheet.cell(row=r,column=c).value)
	return str(worksheet.cell(coor).value)

def getSheetName(sc):

	if sc.find("'") != -1:
		return(sc[sc.find("'")+1:sc.rfind("'")])

	else:
		return val

def array_atize(c):
	i=0
	sc = []
	while i < len(c):
		sc.append(c[i:i+1])
		i+=1

	return sc

def cleanMatrix(pl):

	i = 0
	for each in pl:
		i+=1
		for j in range(i, len(pl)-1):


			if each[1] == pl[j][1]:
				pl= pl[:j-1] + pl[j:]

	return pl

def changeDestinationMatrix(dest, pl):
	spaces = "     "
	dest = dest+spaces[len(dest):]
	i = 0

	for each in pl:
		pl[i][0] = dest + pl[i][0][5:]
		i+=1
	return pl

def pause():
	input("...")

def run():

	#reset matrix
	
	f=open(path +"IMPORT_MATRIX.lsq", 'w')
	f.write("");
	f.close
	
	
	f = open(path + "wirt.lsq", 'r')

	productList = readMatrix(f)

	f.close()
	f = open(path + "wirt2.lsq", 'w')

	productList = changeDestinationMatrix("WIRET", productList)
	writeMatrix(f, productList)


	getcontext().prec = 3
	print("Doing wire...")
	a = do_matrix(WIRE_VENDOR)
	if a == -1:
		return
		
	while True:
		choice = input("Would you like to do pipe as well?")
		if choice.lower() == "y":
			a=do_matrix(PIPE_VENDOR)
				if a == -1:
					return
			break
		elif choice.lower() == 'n':
			break
		else:
			print("Invalid choice.")
	
	print("Done!")

	if __name__ !="__main__":
		input("\nSaved Matrix to 'C:\PaulScripts\Wire Matrix\IMPORT_MATRIX.lsq'\nPress enter to continue...")
		
if __name__ == "__main__":
	run()
else:
	print("imported matrix.py")
